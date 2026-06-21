"""Test ERCOT price loading and SPP parsing.

Validates:
- ERCOT xlsx format parsing (hour/interval conversion)
- Settlement point filtering (LZ_NORTH)
- $/MWh -> $/kWh conversion
- Timezone localization
- Detection of missing settlement points with helpful error
"""
import sys
from pathlib import Path
import tempfile

import pandas as pd
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from txpower.ercot_prices import load_spp_annual_xlsx, synthetic_uri_spp, align_price_to_usage


def _create_ercot_format_xlsx(path: str | Path, num_rows: int = 100, settlement_point: str = "LZ_NORTH") -> None:
    """Create a mock ERCOT RTM SPP xlsx with correct column structure.

    Matches actual ERCOT format:
    - Delivery Date | Delivery Hour | Delivery Interval | Repeated Hour Flag |
      Settlement Point Name | Settlement Point Price
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jan2021"

    # Header row
    ws.append([
        "Delivery Date", "Delivery Hour", "Delivery Interval",
        "Repeated Hour Flag", "Settlement Point Name", "Settlement Point Price"
    ])

    # Data rows: 4 intervals per hour, 24 hours per day
    row_idx = 0
    for day in range(1, 4):  # 3 days of data
        for hour_of_day in range(1, 25):  # 1..24 (hour-ending)
            for interval in range(1, 5):  # 1..4 (four 15-min intervals)
                if row_idx >= num_rows:
                    break
                date_str = f"2021-01-{day:02d}"
                price = 30.0 + row_idx * 0.1  # $/MWh

                ws.append([
                    date_str,           # Delivery Date
                    hour_of_day,        # Delivery Hour (1..24)
                    interval,           # Delivery Interval (1..4)
                    0,                  # Repeated Hour Flag
                    settlement_point,   # Settlement Point Name
                    price,              # Settlement Point Price ($/MWh)
                ])
                row_idx += 1
            if row_idx >= num_rows:
                break
        if row_idx >= num_rows:
            break

    wb.save(path)


def test_load_spp_annual_xlsx_lz_north():
    """Verify LZ_NORTH settlement point is parsed correctly."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        _create_ercot_format_xlsx(path, num_rows=200)
        spp = load_spp_annual_xlsx(path, settlement_point="LZ_NORTH")

        assert isinstance(spp.index, pd.DatetimeIndex)
        assert spp.index.tz is not None
        assert str(spp.index.tz) == "America/Chicago"
        assert len(spp) > 0
        assert all(spp > 0)  # All prices should be positive (in $/kWh)
        assert all(spp < 0.1)  # Max $/MWh 9000 = $9/kWh; test data around $30-50/MWh
        print(f"✓ load_spp_annual_xlsx LZ_NORTH: {len(spp)} prices, "
              f"range ${spp.min():.4f}-${spp.max():.4f}/kWh")
    finally:
        Path(path).unlink()


def test_load_spp_annual_xlsx_conversion_mwh_to_kwh():
    """Verify $/MWh -> $/kWh conversion is correct."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        _create_ercot_format_xlsx(path, num_rows=100)
        spp = load_spp_annual_xlsx(path, settlement_point="LZ_NORTH")

        # Find first few prices in the raw file
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        first_price_mwh = ws["F2"].value
        expected_kwh = first_price_mwh / 1000.0

        # The series should have first price matching the conversion
        first_in_series = spp.iloc[0]
        assert abs(first_in_series - expected_kwh) < 1e-8
        print(f"✓ $/MWh → $/kWh conversion: {first_price_mwh} MWh = ${first_in_series:.6f}/kWh")
    finally:
        Path(path).unlink()


def test_load_spp_annual_xlsx_hour_interval_parsing():
    """Verify Delivery Hour (1..24) and Interval (1..4) are converted correctly."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        _create_ercot_format_xlsx(path, num_rows=100)
        spp = load_spp_annual_xlsx(path, settlement_point="LZ_NORTH")

        # First four prices should be for times 00:00, 00:15, 00:30, 00:45
        expected_times = [
            "2021-01-01 00:00:00",  # Hour 1, Interval 1 -> 0h, 0m
            "2021-01-01 00:15:00",  # Hour 1, Interval 2 -> 0h, 15m
            "2021-01-01 00:30:00",  # Hour 1, Interval 3 -> 0h, 30m
            "2021-01-01 00:45:00",  # Hour 1, Interval 4 -> 0h, 45m
        ]

        for i, expected in enumerate(expected_times):
            expected_ts = pd.Timestamp(expected, tz="America/Chicago")
            actual_ts = spp.index[i]
            assert actual_ts == expected_ts, f"Index {i}: expected {expected_ts}, got {actual_ts}"

        print(f"✓ Hour/Interval parsing: times aligned correctly at 15-min resolution")
    finally:
        Path(path).unlink()


def test_load_spp_annual_xlsx_missing_settlement_point():
    """Verify helpful error when settlement point is not found."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        # Create xlsx with only LZ_SOUTH
        _create_ercot_format_xlsx(path, num_rows=50, settlement_point="LZ_SOUTH")
        try:
            load_spp_annual_xlsx(path, settlement_point="NONEXISTENT_ZONE")
            assert False, "Should have raised ValueError"
        except ValueError as e:
            assert "NONEXISTENT_ZONE" in str(e)
            assert "LZ_SOUTH" in str(e)
            print(f"✓ Missing settlement point error: detected {e}")
    finally:
        Path(path).unlink()


def test_synthetic_uri_spp_structure():
    """Verify synthetic Uri SPP has expected structure."""
    spp = synthetic_uri_spp()

    assert isinstance(spp.index, pd.DatetimeIndex)
    assert spp.index.tz is not None
    assert len(spp) > 0

    # Check that synthetic data is clearly marked
    assert spp.attrs.get("synthetic") is True

    # Feb 15-19 should be at the price cap
    cap_mask = (spp.index.strftime("%m-%d") >= "02-15") & (spp.index.strftime("%m-%d") <= "02-19")
    cap_prices = spp[cap_mask]
    assert all(abs(cap_prices - 9.0) < 1e-6)  # $9/kWh = $9000/MWh cap

    # Core month (early Feb) should be low
    core_mask = (spp.index.strftime("%m-%d") >= "02-01") & (spp.index.strftime("%m-%d") < "02-14")
    core_prices = spp[core_mask]
    assert all(core_prices < 0.1)  # Below $100/MWh

    # Shoulders may be elevated ($500-$6000/MWh = $0.5-$6/kWh)
    assert spp.min() < 0.1  # Some very cheap prices (normal periods)

    print(f"✓ Synthetic Uri SPP: {len(cap_prices)} capped intervals, "
          f"core month range ${core_prices.min():.4f}-${core_prices.max():.4f}/kWh, "
          f"overall min ${spp.min():.4f}/kWh")


def test_align_price_to_usage():
    """Verify price alignment resamples correctly to finer usage grid."""
    spp_15min = synthetic_uri_spp()

    # Create a 1-minute usage index
    usage_idx = pd.date_range("2021-02-01", periods=1440, freq="1min", tz="America/Chicago")

    aligned = align_price_to_usage(spp_15min, usage_idx)

    # Should have same length as usage index
    assert len(aligned) == len(usage_idx)

    # Each 15-min price should be repeated across 15 1-min intervals
    assert aligned.index[0] == usage_idx[0]
    assert aligned.iloc[0] == aligned.iloc[14]  # First 15 values should be same
    assert aligned.iloc[15] != aligned.iloc[14]  # 16th value should differ (next 15-min bucket)

    print(f"✓ Align price to usage: {len(spp_15min)} 15-min prices → {len(aligned)} 1-min with ffill")


if __name__ == "__main__":
    test_load_spp_annual_xlsx_lz_north()
    test_load_spp_annual_xlsx_conversion_mwh_to_kwh()
    test_load_spp_annual_xlsx_hour_interval_parsing()
    test_load_spp_annual_xlsx_missing_settlement_point()
    test_synthetic_uri_spp_structure()
    test_align_price_to_usage()
    print("\n✓ All ERCOT price tests passed.")
